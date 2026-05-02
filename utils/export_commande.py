import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

def export_commande_excel(df_commande, nom_societe="Entrepôt Central", pays_fournisseur="Portugal"):
    """
    Génère un fichier Excel professionnel pour la commande fournisseur.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Commande Fournisseur"

    # Couleurs
    bleu_fonce = PatternFill("solid", fgColor="1F77B4")
    bleu_clair = PatternFill("solid", fgColor="D6E4F0")
    rouge = PatternFill("solid", fgColor="FF4444")
    orange = PatternFill("solid", fgColor="FF8C00")
    jaune = PatternFill("solid", fgColor="FFD700")
    vert = PatternFill("solid", fgColor="90EE90")
    gris = PatternFill("solid", fgColor="F2F2F2")

    blanc_gras = Font(bold=True, color="FFFFFF", size=12)
    noir_gras = Font(bold=True, size=11)
    normal = Font(size=10)

    # Bordures
    bordure = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ---------------------------------------------------------
    # EN-TÊTE DU DOCUMENT
    # ---------------------------------------------------------
    ws.merge_cells("A1:G1")
    ws["A1"] = f"BON DE COMMANDE FOURNISSEUR — {pays_fournisseur.upper()}"
    ws["A1"].font = Font(bold=True, size=16, color="1F77B4")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:G2")
    ws["A2"] = f"{nom_societe}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Date de commande : {datetime.now().strftime('%d/%m/%Y')}  |  Délai livraison estimé : 3 à 4 semaines"
    ws["A3"].font = Font(italic=True, size=10, color="666666")
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[4].height = 10  # Ligne vide

    # ---------------------------------------------------------
    # RÉSUMÉ
    # ---------------------------------------------------------
    ws["A5"] = "RÉSUMÉ DE LA COMMANDE"
    ws["A5"].font = noir_gras
    ws["A5"].fill = bleu_clair

    ws["A6"] = "Nombre de références :"
    ws["B6"] = len(df_commande)
    ws["B6"].font = noir_gras

    ws["D6"] = "Dont CRITIQUES :"
    critiques = len(df_commande[df_commande["statut"].str.contains("CRITIQUE", na=False)])
    ws["E6"] = critiques
    ws["E6"].font = Font(bold=True, color="FF4444")

    ws["D7"] = "Dont À COMMANDER :"
    a_commander = len(df_commande[df_commande["statut"].str.contains("COMMANDER", na=False)])
    ws["E7"] = a_commander
    ws["E7"].font = Font(bold=True, color="FF8C00")

    ws.row_dimensions[8].height = 10  # Ligne vide

    # ---------------------------------------------------------
    # EN-TÊTE DU TABLEAU
    # ---------------------------------------------------------
    headers = [
        "Référence",
        "Désignation",
        "Stock Actuel",
        "Point de Commande",
        "Qté à Commander",
        "Statut",
        "Observations"
    ]

    row_header = 9
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col, value=header)
        cell.fill = bleu_fonce
        cell.font = blanc_gras
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bordure

    ws.row_dimensions[row_header].height = 25

    # ---------------------------------------------------------
    # DONNÉES
    # ---------------------------------------------------------
    for row_idx, (_, row) in enumerate(df_commande.iterrows(), row_header + 1):
        statut = str(row.get("statut", ""))

        # Couleur selon statut
        if "RUPTURE" in statut:
            fill_row = rouge
        elif "CRITIQUE" in statut:
            fill_row = orange
        elif "COMMANDER" in statut:
            fill_row = jaune
        else:
            fill_row = gris if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")

        valeurs = [
            row.get("reference", ""),
            row.get("designation", ""),
            row.get("quantite_stock", 0),
            row.get("point_commande", 0),
            row.get("quantite_a_commander", 0),
            statut,
            ""  # Observations vide
        ]

        for col, valeur in enumerate(valeurs, 1):
            cell = ws.cell(row=row_idx, column=col, value=valeur)
            cell.fill = fill_row
            cell.font = normal
            cell.border = bordure
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                        vertical="center")

        ws.row_dimensions[row_idx].height = 20

    # ---------------------------------------------------------
    # LARGEURS DES COLONNES
    # ---------------------------------------------------------
    largeurs = [15, 30, 15, 20, 18, 18, 25]
    for col, largeur in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(col)].width = largeur

    # ---------------------------------------------------------
    # PIED DE PAGE
    # ---------------------------------------------------------
    last_row = row_header + len(df_commande) + 2
    ws.merge_cells(f"A{last_row}:G{last_row}")
    ws[f"A{last_row}"] = "Document généré par Superviseur Stock & Approvisionnement — SMD Consulting"
    ws[f"A{last_row}"].font = Font(italic=True, size=9, color="999999")
    ws[f"A{last_row}"].alignment = Alignment(horizontal="center")

    # ---------------------------------------------------------
    # LÉGENDE
    # ---------------------------------------------------------
    last_row += 2
    ws[f"A{last_row}"] = "LÉGENDE :"
    ws[f"A{last_row}"].font = noir_gras

    legendes = [
        (last_row + 1, rouge, "🔴 RUPTURE — Stock épuisé, commande urgente"),
        (last_row + 2, orange, "🟠 CRITIQUE — Stock en dessous du minimum"),
        (last_row + 3, jaune, "🟡 À COMMANDER — Stock en dessous du point de commande"),
    ]

    for row_l, fill_l, texte_l in legendes:
        ws.merge_cells(f"A{row_l}:G{row_l}")
        ws[f"A{row_l}"] = texte_l
        ws[f"A{row_l}"].fill = fill_l
        ws[f"A{row_l}"].font = Font(size=9)

    # Sauvegarde en mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer